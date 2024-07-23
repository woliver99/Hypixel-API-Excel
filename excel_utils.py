import openpyxl

excel_file = "investments.xlsx"

def getWorkbook():
    return openpyxl.load_workbook(excel_file)

def getItemNames(sheet: str) -> set[str]:
    workbook = getWorkbook()
    sheet = workbook[sheet]

    items = set()
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1, values_only=True):
        items.add(row[0])

    if None in items:
        items.remove(None)
        
    return items

def updateItemPrices(sheet:str, prices: dict[str, int]):
    workbook = getWorkbook()
    sheet = workbook[sheet]

    for row in range(2, sheet.max_row + 1):
        item = sheet.cell(row=row, column=1).value
        if item in prices:
            sheet.cell(row=row, column=2).value = prices[item]

    try:
        workbook.save(excel_file)
    except Exception as e:
        print("Error saving workbook, is it already opened?: ", e)
        input("Press enter to close...")
        exit()